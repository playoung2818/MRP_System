from __future__ import annotations

import json
import os
import tempfile
import uuid
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import text

from erp_system.runtime.db_config import get_engine
from erp_system.runtime.policies import GOOGLE_SHEET_SPREADSHEET, GOOGLE_SHEET_WORKSHEET

from ._helpers import (
    ServiceAccountCredentials,
    _copy_via_powershell,
    _reset_gsheet_user_format,
    _resolve_google_cred_path,
    gspread,
    set_with_dataframe,
)


def engine():
    return get_engine()


def read_table_if_exists(schema: str, table: str) -> pd.DataFrame:
    eng = engine()
    query = f'SELECT * FROM "{schema}"."{table}"'
    try:
        return pd.read_sql(query, eng)
    except Exception:
        return pd.DataFrame()


def write_to_db(df: pd.DataFrame, schema: str, table: str):
    if df is None:
        return
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == "object":
            out[c] = out[c].map(
                lambda v: json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list, tuple, set)) else v
            )
    out = out.where(pd.notna(out), None)
    eng = engine()
    try:
        with eng.begin() as conn:
            conn.execute(text(f'CREATE SCHEMA IF NOT EXISTS "{schema}"'))
    except Exception:
        pass
    try:
        out.to_sql(table, eng, schema=schema, if_exists="replace", index=False, method="multi", chunksize=2000)
    except Exception as exc:
        msg = str(exc)
        if ("sqlalche.me/e/20/e3q8" in msg) or ("bind parameter" in msg.lower()):
            out.to_sql(table, eng, schema=schema, if_exists="replace", index=False, method=None, chunksize=500)
        else:
            raise


def write_final_sales_order_to_gsheet(
    df: pd.DataFrame,
    *,
    spreadsheet_name: str = GOOGLE_SHEET_SPREADSHEET,
    worksheet_name: str = GOOGLE_SHEET_WORKSHEET,
    cred_path: str | None = None,
):
    if gspread is None or set_with_dataframe is None or ServiceAccountCredentials is None:
        raise ImportError("Google Sheets dependencies are missing. Install: pip install gspread gspread-dataframe oauth2client")
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    resolved_cred_path = _resolve_google_cred_path(cred_path)
    cred_path_use = resolved_cred_path
    temp_cred_path = None
    try:
        temp_cred_path = os.path.join(tempfile.gettempdir(), f"gsheet_cred_{uuid.uuid4().hex}.json")
        _copy_via_powershell(resolved_cred_path, temp_cred_path)
        cred_path_use = temp_cred_path
    except Exception:
        cred_path_use = resolved_cred_path
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(cred_path_use, scope)
        client = gspread.authorize(creds)
        sh = client.open(spreadsheet_name)
        export_df = df.copy()
        if "SO Entry Date" in export_df.columns:
            export_df = export_df.drop(columns=["SO Entry Date"])
        if "Remark" in export_df.columns:
            export_df = export_df.drop(columns=["Remark"]) # Drop old Remark completely
        remarks = pd.Series("", index=export_df.index, dtype="string")
        if "QB Num" in export_df.columns:
            try:
                remark_ws = sh.worksheet("SO_Remark")
                remark_rows = pd.DataFrame(remark_ws.get_all_records())
                if not remark_rows.empty and "QB Num" in remark_rows.columns and "Remark" in remark_rows.columns:
                    remark_rows["QB Num"] = remark_rows["QB Num"].astype(str).str.strip()
                    remark_rows["Remark"] = remark_rows["Remark"].astype(str).str.strip()
                    remark_map = (
                        remark_rows.loc[remark_rows["QB Num"].ne(""), ["QB Num", "Remark"]] # filter rows where QB Num is not Blank and only keep ["QB Num", "Remark"]
                        .drop_duplicates(subset=["QB Num"], keep="last")
                        .set_index("QB Num")["Remark"]
                    )
                    remarks = export_df["QB Num"].astype(str).str.strip().map(remark_map).fillna("")
            except Exception:
                remarks = pd.Series("", index=export_df.index, dtype="string")
        export_df.insert(0, "Remark", remarks)
        try:
            ws = sh.worksheet(worksheet_name)
            ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=worksheet_name, rows=100, cols=26)
        _reset_gsheet_user_format(ws)
        set_with_dataframe(ws, export_df, include_index=False, include_column_header=True, resize=True)
        try:
            ws.freeze(rows=1)
        except Exception:
            pass
        print(f"Wrote {len(export_df)} rows to Google Sheet -> {spreadsheet_name} / {worksheet_name}")
    finally:
        if temp_cred_path:
            try:
                if os.path.exists(temp_cred_path):
                    os.remove(temp_cred_path)
            except Exception:
                pass


def save_not_assigned_so(
    df: pd.DataFrame,
    output_path: str = "Not_assigned_SO.xlsx",
    highlight_cols: str | list[str] | None = None,
    band_by_col: str = "QB Num",
    shortage_col: str = "Component_Status",
    shortage_value: str = "Shortage",
    column_widths: dict | None = None,
    pod_watchlist_df: pd.DataFrame | None = None,
    pod_watchlist_sheet: str = "POD-Wachlist",
) -> dict:
    export_columns = [
        "Order Date",
        "Name",
        "QB Num",
        "Item",
        "Qty(-)",
        "Available",
        "Available + On PO",
        "Sales/Week",
        "Recommended Restock Qty",
        "Assigned Q'ty",
        "On Hand",
        "On Sales Order",
        "On PO",
        "Component_Status",
        # "P. O. #",
        # "Ship Date",
    ]
    export_df = df[[col for col in export_columns if col in df.columns]].copy()
    date_columns = ["Order Date", "Ship Date"]
    for col in date_columns:
        if col in export_df.columns:
            export_df[col] = pd.to_datetime(export_df[col], errors="coerce").dt.floor("D")

    if column_widths is None:
        column_widths = {
            "Order Date": 15,
            "Item": 30,
            "Name": 25,
            "P. O. #": 5,
            "QB Num": 15,
            "Qty(-)": 10,
            "Available": 15,
            "Available + Pre-installed PO": 25,
            "On Hand - WIP": 17,
            "Recommended Restock Qty": 25,
            "On Sales Order": 15,
            "Available + On PO": 20,
            "Assigned Q'ty": 15,
            "Sales/Week": 15,
        }

    output_path = str(output_path)
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    use_xlsx_fallback = output_path.lower().endswith(".xlsx")

    if use_xlsx_fallback:
        wb = None
        if os.path.exists(output_path):
            temp_in = os.path.join(tempfile.gettempdir(), f"not_assigned_in_{uuid.uuid4().hex}.zip")
            try:
                _copy_via_powershell(output_path, temp_in)
                with open(temp_in, "rb") as f:
                    wb = load_workbook(BytesIO(f.read()))
            except Exception:
                wb = None
            finally:
                try:
                    if os.path.exists(temp_in):
                        os.remove(temp_in)
                except Exception:
                    pass
        if wb is None:
            wb = Workbook()
        first_sheet_name = wb.sheetnames[0] if wb.sheetnames else "Sheet1"
        if first_sheet_name in wb.sheetnames:
            del wb[first_sheet_name]
        ws = wb.create_sheet(title=first_sheet_name, index=0)
        for row in dataframe_to_rows(export_df, index=False, header=True):
            ws.append(row)
        if pod_watchlist_df is not None:
            if pod_watchlist_sheet in wb.sheetnames:
                del wb[pod_watchlist_sheet]
            ws_pod = wb.create_sheet(title=pod_watchlist_sheet)
            for row in dataframe_to_rows(pod_watchlist_df, index=False, header=True):
                ws_pod.append(row)
    else:
        if not os.path.exists(output_path):
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                export_df.to_excel(writer, sheet_name="Sheet1", index=False)

        _wb = load_workbook(output_path)
        first_sheet_name = _wb.sheetnames[0]
        _wb.close()

        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            export_df.to_excel(writer, sheet_name=first_sheet_name, index=False)
            if pod_watchlist_df is not None:
                pod_watchlist_df.to_excel(writer, sheet_name=pod_watchlist_sheet, index=False)
        wb = load_workbook(output_path)
        ws = wb.worksheets[0]

    ws.freeze_panes = "A2"
    col_map: dict[str, int] = {}
    band_col_idx = None
    shortage_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        header = cell.value
        col_map[header] = idx
        if header == band_by_col:
            band_col_idx = idx
        if header == shortage_col:
            shortage_col_idx = idx

    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="00FF0000")
    center_align = Alignment(horizontal="center", vertical="center")

    if band_col_idx is not None and shortage_col_idx is not None:
        current_key = None
        fill_toggle = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            band_cell = row[band_col_idx - 1]
            status_cell = row[shortage_col_idx - 1]
            if band_cell.value != current_key:
                current_key = band_cell.value
                fill_toggle = not fill_toggle
            row_fill = gray_fill if fill_toggle else white_fill
            for c in row:
                c.fill = row_fill
            if status_cell.value == shortage_value:
                for c in row:
                    c.font = red_font

    sales_idx = col_map.get("Sales/Week")
    avail_on_po_idx = col_map.get("Available + On PO")
    if sales_idx and avail_on_po_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            try:
                sales_val = float(row[sales_idx - 1].value)
            except (TypeError, ValueError):
                sales_val = 0.0
            try:
                avail_val = float(row[avail_on_po_idx - 1].value)
            except (TypeError, ValueError):
                avail_val = 0.0
            if sales_val > avail_val:
                row[avail_on_po_idx - 1].fill = yellow_fill

    if highlight_cols is None:
        highlight_cols = []
    if isinstance(highlight_cols, str):
        highlight_cols = [highlight_cols]
    for highlight_col in highlight_cols:
        if highlight_col in col_map:
            h_idx = col_map[highlight_col]
            for (cell,) in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=h_idx, max_col=h_idx):
                try:
                    val = float(cell.value)
                except (TypeError, ValueError):
                    val = 0.0
                if val > 0:
                    cell.fill = yellow_fill

    for name, width in column_widths.items():
        if name in col_map:
            letter = ws.cell(row=1, column=col_map[name]).column_letter
            ws.column_dimensions[letter].width = width

    for name in ["Qty(-)", "Available + Pre-installed PO", "Available", "Available + On PO", "Sales/Week", "Recommended Restock Qty"]:
        if name in col_map:
            idx = col_map[name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    cell.alignment = center_align

    for name in date_columns:
        if name in col_map:
            idx = col_map[name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    if cell.value not in (None, ""):
                        cell.number_format = "yyyy-mm-dd"

    today_str = datetime.today().strftime("%Y-%m-%d")
    ws.title = today_str
    if pod_watchlist_df is not None and pod_watchlist_sheet in wb.sheetnames:
        wb[pod_watchlist_sheet].freeze_panes = "A2"

    if use_xlsx_fallback:
        out_bytes = BytesIO()
        wb.save(out_bytes)
        temp_out = os.path.join(tempfile.gettempdir(), f"not_assigned_out_{uuid.uuid4().hex}.zip")
        with open(temp_out, "wb") as f:
            f.write(out_bytes.getvalue())
        try:
            _copy_via_powershell(temp_out, output_path)
        finally:
            try:
                if os.path.exists(temp_out):
                    os.remove(temp_out)
            except Exception:
                pass
    else:
        wb.save(output_path)

    unique_wo = df[band_by_col].nunique() if band_by_col in df.columns else 0
    return {"Number of unassigned WOs:": unique_wo, "sheet_name": today_str}


__all__ = [
    "read_table_if_exists",
    "save_not_assigned_so",
    "write_final_sales_order_to_gsheet",
    "write_to_db",
]
