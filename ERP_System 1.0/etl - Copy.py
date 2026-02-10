import json, re, numpy as np, pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime
from gspread_dataframe import set_with_dataframe
from oauth2client.service_account import ServiceAccountCredentials # type: ignore
import gspread
import logging
from flask import Flask, jsonify 
from flask_sqlalchemy import SQLAlchemy  # type: ignore
from config import ( # type: ignore
    SALES_ORDER_FILE, WAREHOUSE_INV_FILE, SHIPPING_SCHEDULE_FILE, POD_FILE,
    DATABASE_DSN, DB_SCHEMA, TBL_INVENTORY, TBL_STRUCTURED, TBL_SALES_ORDER, TBL_POD, TBL_Shipping
)
import logging, sys, traceback
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook
import os
from pathlib import Path

ERP_MODULE_DIR = Path(__file__).resolve().parents[1] / "ERP_System 2.0"
if str(ERP_MODULE_DIR) not in sys.path:
    sys.path.append(str(ERP_MODULE_DIR))

from erp_normalize import normalize_item



# --------------------
# Helpers
# --------------------
def engine():
    return create_engine(DATABASE_DSN, pool_pre_ping=True)

def write_inventory_status(df: pd.DataFrame, schema: str = "public", table: str = "inventory_status"):
    eng = engine()
    df.to_sql(
        table,                 
        eng,
        schema=schema,
        if_exists="replace",   # drops + recreates to match df columns
        index=False,
        method="multi",
        chunksize=10_000,
    )

def write_sales_order(df_sales_order: pd.DataFrame, schema: str = "public", table: str = "sales_order"):
    eng = engine()
    df_sales_order.to_sql(
    table,                 
    eng,
    schema=schema,
    if_exists="replace",   # drops + recreates to match df columns
    index=False,
    method="multi",
    chunksize=10_000,
)

def write_structured(structured_df: pd.DataFrame, schema: str = "public", table: str = "wo_structured"):
    desired_order = [
    "Order Date", "Name", "P. O. #", "QB Num", "Item", "Qty",
    "Available + Pre-installed PO", "Available", "On Hand", "On Hand - WIP",
    "Assigned Q'ty", "On Sales Order", "On PO",
    "Available + On PO", "Sales/Week", "Recommended Restock Qty",
    "Ship Date", "Picked", "SO_Status", "Qty(+)"
]

    ordered = [c for c in desired_order if c in structured_df.columns] + \
            [c for c in structured_df.columns if c not in desired_order]
    structured_df = structured_df.reindex(columns=ordered)

    eng = engine()
    structured_df.to_sql(
        table,
        eng,
        schema=schema,
        if_exists="replace",
        index=False,
        method="multi",
        chunksize=10_000,
    )

def write_pod(df_pod: pd.DataFrame, schema: str = "public", table: str = "open_purchase_order"):
    eng = engine()
    df_pod.to_sql(
        table,
        eng,
        schema=schema,
        if_exists="replace",   # drops + recreates to match df
        index=False,
        method="multi",
        chunksize=10_000,
    )

def write_Shipping_Schedule(df_shipping_schedule : pd.DataFrame, schema: str = "public", table: str = "NT Shipping Schedule"):
    eng = engine()
    df_shipping_schedule .to_sql(
        table,
        eng,
        schema=schema,
        if_exists="replace",   # drops + recreates to match df
        index=False,
        method="multi",
        chunksize=10_000,
    )

def write_final_sales_order_to_gsheet(
    df: pd.DataFrame,
    *,
    spreadsheet_name: str,
    worksheet_name: str,
    cred_path: str,
):
    """
    Overwrites (or creates) the given worksheet and writes df with headers.
    Auto-resizes the sheet to fit df shape.
    """
    # 1) auth
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name(cred_path, scope)
    client = gspread.authorize(creds)

    # 2) open spreadsheet (by name) — or use client.open_by_key("<spreadsheet_id>")
    sh = client.open(spreadsheet_name)

    # 3) get/create worksheet
    try:
        ws = sh.worksheet(worksheet_name)
        ws.clear()
    except gspread.exceptions.WorksheetNotFound:
        # at least 100 rows/cols to avoid size errors; we’ll resize next
        ws = sh.add_worksheet(title=worksheet_name, rows=100, cols=26)

    # 4) write
    set_with_dataframe(ws, df, include_index=False, include_column_header=True, resize=True)

    # freeze header row for readability
    try:
        ws.freeze(rows=1)
    except Exception:
        pass

    print(f"✅ Wrote {len(df)} rows to Google Sheet → {spreadsheet_name} / {worksheet_name}")

def normalize_wo_number(wo: str) -> str:
    m = re.search(r'\b(20\d{6})\b', str(wo))
    return f"SO-{m.group(1)}" if m else str(wo)

# --------------------
# Extract
# --------------------
def extract_inputs():
    df_sales_order = pd.read_csv(SALES_ORDER_FILE, encoding="ISO-8859-1")
    inventory_df = pd.read_csv(WAREHOUSE_INV_FILE)
    df_shipping_schedule = pd.read_excel(SHIPPING_SCHEDULE_FILE)
    df_pod = pd.read_csv(POD_FILE, encoding="ISO-8859-1")
    return df_sales_order, inventory_df, df_shipping_schedule, df_pod

def fetch_word_files_df(api_url: str) -> pd.DataFrame:
    import requests
    try:
        r = requests.get(api_url, timeout=10)
        r.raise_for_status()
        data = r.json()
        wf = pd.DataFrame(data.get("word_files", []))
    except Exception:
        wf = pd.DataFrame(columns=["file_name","order_id","status"])
    if "order_id" in wf.columns:
        wf = wf.rename(columns={"order_id":"WO_Number"})
    wf["WO_Number"] = wf["WO_Number"].astype(str).apply(normalize_wo_number)
    return wf


def fetch_pdf_orders_df_from_supabase(dsn: str) -> pd.DataFrame:
    """
    Reads order_id + extracted_data from public.pdf_file_log and returns a
    two-column DataFrame with ['WO','Product Number'] rows, one per item in JSON.
    """
    eng = create_engine(dsn, pool_pre_ping=True)
    rows = pd.read_sql('SELECT order_id, extracted_data FROM public.pdf_file_log', eng)

    def rows_from_json(extracted_data, order_id=""):
        # extracted_data may be JSON string or dict
        if isinstance(extracted_data, str):
            try:
                extracted_data = json.loads(extracted_data)
            except Exception:
                extracted_data = {}
        data = extracted_data or {}
        wo = data.get("wo", order_id)
        items = data.get("items") or []

        # one row per item; if none, keep a placeholder
        if not items:
            return [{"WO": wo, "Product Number": ""}]
        out = []
        for it in items:
            pn = (
                it.get("product_number")
                or it.get("part_number")
                or it.get("product")
                or it.get("part")
                or ""
            )
            out.append({"WO": wo, "Product Number": pn})
        return out

    all_rows = []
    for _, r in rows.iterrows():
        all_rows.extend(rows_from_json(r.get("extracted_data"), r.get("order_id")))

    return pd.DataFrame(all_rows, columns=["WO", "Product Number"])


# --------------------
# Transform 
# --------------------
def transform_sales_order(df_sales_order: pd.DataFrame) -> pd.DataFrame:
    df = df_sales_order.copy()
    df = df.drop(columns = ['Qty', 'Item'])
    df = df.rename(columns={"Unnamed: 0": "Item", "Num": "QB Num", "Backordered": "Qty(-)", "Date":"Order Date"})
    df["Item"] = df["Item"].ffill().astype(str).str.strip()
    df = df[~df["Item"].str.startswith("total", na=False)]
    df = df[~df["Item"].str.lower().isin(["forwarding charge", "tariff (estimation)"])]
    df = df[df["Inventory Site"] == "WH01S-NTA"]
    df["Item"] = df["Item"].map(normalize_item)
    return df

def transform_inventory(inventory_df: pd.DataFrame) -> pd.DataFrame:
    inv = inventory_df.copy()
    # only rename ONCE
    inv = inv.rename(columns={"Unnamed: 0":"Part_Number"})
    inv["Part_Number"] = inv["Part_Number"].astype(str).str.strip()
    inv["Part_Number"] = inv["Part_Number"].map(normalize_item)
    # make numeric safely
    for c in ["On Hand","On Sales Order","On PO","Available"]:
        if c in inv.columns:
            inv[c] = pd.to_numeric(inv[c], errors="coerce").fillna(0)
    return inv

def transform_pod(df_pod: pd.DataFrame) -> pd.DataFrame:
    pod = df_pod.drop(columns=['Amount','Open Balance',"Rcv'd","Qty"], axis =1)
    pod.rename(columns={"Date":"Order Date","Num":"QB Num","Backordered":"Qty(+)"},inplace=True)
    pod = pod.drop(pod.columns[[0]], axis =1)
    pod = pod.dropna(axis=0, how='all',subset=None, inplace=False)
    pod = pod.dropna(thresh=5)
    pod['Memo'] = pod['Memo'].str.split(' ',expand=True)[0]
    pod['QB Num'] = pod['QB Num'].str.split('(',expand=True)[0]
    # print(pod['Memo'].str.split('*',expand=True)[0])
    pod['Memo'] = pod['Memo'].str.replace("*","")
    pod.rename(columns={"Memo":"Item"},inplace=True)
    pod['Order Date']= pd.to_datetime(pod['Order Date'])
    pod['Deliv Date']= pd.to_datetime(pod['Deliv Date'])
    pod['Order Date'] = pod['Order Date'].dt.strftime('%Y/%m/%d')
    pod['Deliv Date'] = pod['Deliv Date'].dt.strftime('%Y/%m/%d')
    pod["Item"] = pod["Item"].map(normalize_item)
    df_pod = pd.DataFrame(pod)
    return df_pod


def transform_shipping(df_shipping_schedule: pd.DataFrame) -> pd.DataFrame:

    df = df_shipping_schedule.copy()

    # --- make sure the columns exist (create empty ones if missing) ---
    need = ['SO NO.', 'Customer PO No.', 'Model Name', 'Ship Date', 'Qty', 'Description']
    for c in need:
        if c not in df.columns:
            df[c] = np.nan

    # --- select and rename ---
    Ship = df.loc[:, need].copy()
    Ship.rename(columns={
        "Customer PO No.": "QB Num",
        "Model Name": "Item",
        "Qty": "Qty(+)"
    }, inplace=True)

    # --- basic cleaning ---
    # QB Num: strip anything after '('
    Ship["QB Num"] = Ship["QB Num"].astype(str).str.split("(").str[0].str.strip()

    # types
    Ship["Item"] = Ship["Item"].astype(str).str.strip()
    Ship["Description"] = Ship["Description"].astype(str)

    # coerce Ship Date to yyyy/mm/dd string if you want it normalized (optional)
    Ship["Ship Date"] = pd.to_datetime(Ship["Ship Date"], errors="coerce").dt.date

    # Qty(+) numeric
    Ship["Qty(+)"] = pd.to_numeric(Ship["Qty(+)"], errors="coerce").fillna(0).astype(int)

    # --- Pre/Bare logic ---
    model_ok = Ship["Item"].str.upper().str.startswith(("N", "SEMIL", "POC"), na=False)
    # accept English or Chinese comma: ", including" or "， including"
    including_ok = Ship["Description"].str.contains(r"[，,]\s*including\b", case=False, na=False)

    pre_mask = model_ok & including_ok
    Ship["Pre/Bare"] = np.where(pre_mask, "Pre", "Bare")

    # optional: tidy column order
    desired = ["SO NO.", "QB Num", "Item", "Description", "Ship Date", "Qty(+)", "Pre/Bare"]
    Ship = Ship.reindex(columns=[c for c in desired if c in Ship.columns] +
                               [c for c in Ship.columns if c not in desired])

    return Ship

def reorder_df_out_by_output(output_df: pd.DataFrame, df_out: pd.DataFrame) -> pd.DataFrame:
    """
    Reorder df_out to match the line ordering found in output_df.
    Both frames are expected to use columns: ['QB Num', 'Item'].
    """
    ref = output_df.copy()
    ref['__pos_out'] = ref.groupby('QB Num').cumcount()              # position within QB Num
    ref['__occ'] = ref.groupby(['QB Num', 'Item']).cumcount()        # occurrence index per (QB Num, Item)
    ref_key = ref[['QB Num', 'Item', '__occ', '__pos_out']]

    tgt = df_out.copy()
    tgt['__occ'] = tgt.groupby(['QB Num', 'Item']).cumcount()

    merged = tgt.merge(ref_key, on=['QB Num', 'Item', '__occ'], how='left')

    merged['__fallback'] = merged.groupby('QB Num').cumcount()
    merged['__pos_out'] = merged['__pos_out'].fillna(np.inf)

    ordered = (
        merged.sort_values(['QB Num', '__pos_out', '__fallback'])
              .drop(columns=['__occ', '__pos_out', '__fallback'])
              .reset_index(drop=True)
    )
    return ordered



def extracted_to_df(order):
    """Return a 2‑column DataFrame: ['WO','Product Number'] for one PDFFileLog row."""
    if order is None:
        return pd.DataFrame(columns=["WO", "Product Number"])

    data = order['extracted_data'] or {}
    if isinstance(data, str):              # if stored as TEXT/JSONB string
        try:
            data = json.loads(data)
        except Exception:
            data = {}

    items = data.get("items") or []
    wo = data.get("wo") or getattr(order, "order_id", "")

    rows = [{
        "WO": wo,
        "Product Number": (
            it.get("product_number") or it.get("part_number")
            or it.get("product") or it.get("part") or ""
        ),
    } for it in items]

    if not rows:
        rows = [{"WO": wo, "Product Number": ""}]

    return pd.DataFrame(rows, columns=["WO", "Product Number"])

def _norm_dash_series(s: pd.Series) -> pd.Series:
    # unify fancy dashes to ASCII hyphen
    return s.str.replace(r"[\u2012\u2013\u2014\u2212]", "-", regex=True)

# def normalize_for_match(df: pd.DataFrame) -> pd.DataFrame:
#     """Normalize WO + Product Number for matching (trim, dashes, case)."""
#     out = df.copy()
#     out["WO"] = out["WO"].astype(str).str.strip().apply(normalize_wo_number)
#     out["Product Number"] = (
#         out["Product Number"].astype(str).str.strip().pipe(_norm_dash_series)
#     )
#     return out

def enforce_column_order(df: pd.DataFrame, order: list[str]) -> pd.DataFrame:
    """Reorder columns to `order`, keeping any extras at the end."""
    front = [c for c in order if c in df.columns]
    back  = [c for c in df.columns if c not in front]
    return df.loc[:, front + back]


def build_structured_df(
    df_sales_order: pd.DataFrame,
    word_files_df: pd.DataFrame,
    inventory_df: pd.DataFrame,
    pdf_orders_df: pd.DataFrame,
    df_pod: pd.DataFrame
) -> pd.DataFrame:

    # Build df_out from Sales Order (standardize to these final names)
    needed_cols = {
        "Order Date": "SO Entry Date",
        "Name": "Customer",
        "P. O. #": "Customer PO",
        "QB Num": "QB Num",
        "Item": "Item",                # <- part key
        "Qty(-)": "Qty",               # <- create Qty from Backordered/Qty(-)
        "Ship Date": "Lead Time"       # <- keep the name used later
    }
    for src in list(needed_cols.keys()):
        if src not in df_sales_order.columns:
            df_sales_order[src] = "" if src not in ("Qty(-)",) else 0

    df_out = df_sales_order.rename(columns=needed_cols)[list(needed_cols.values())].copy()

    # Keep an auxiliary WO column if source has it (for 'Picked' merge fallback)
    for alt in ["WO", "WO_Number", "NTA Order ID", "SO Number"]:
        if alt in df_sales_order.columns:
            df_out["WO"] = df_sales_order[alt].astype(str).apply(normalize_wo_number)
            break
    if "WO" not in df_out.columns:
        df_out["WO"] = ""

    # Sort to group visually
    df_out = df_out.sort_values(['QB Num', 'Item']).reset_index(drop=True)

    # Rename the PDF reference to match our new keys
    pdf_ref = pdf_orders_df.rename(columns={'WO': 'QB Num', 'Product Number': 'Item'})
    final_sales_order = reorder_df_out_by_output(pdf_ref, df_out)

    # Map the short->long part names
    final_sales_order["Item"] = final_sales_order["Item"].map(normalize_item)
    final_sales_order = final_sales_order.loc[:, ~final_sales_order.columns.duplicated()]


    # Merge “Picked” status (collapse per order key)
    word_pick = word_files_df.copy()
    key_used = None

    if "QB Num" in word_pick.columns:
        key_used = "QB Num"
    elif "WO_Number" in word_pick.columns:
        key_used = "WO_Number"
        # normalize to SO-######## format
        word_pick["WO_Number"] = word_pick["WO_Number"].astype(str).apply(normalize_wo_number)

    word_pick["Picked"] = word_pick["status"].astype(str).str.strip().eq("Picked")
    word_pick = word_pick.groupby(key_used, as_index=False)["Picked"].max() if key_used else pd.DataFrame(columns=["WO","Picked"])

    if key_used == "QB Num":
        df_Order_Picked = final_sales_order.merge(word_pick, on="QB Num", how="left")
    elif key_used == "WO_Number":
        df_Order_Picked = final_sales_order.merge(word_pick, left_on="QB Num", right_on="WO_Number", how="left").drop(columns=["WO_Number"])
    else:
        df_Order_Picked = final_sales_order.copy()
        df_Order_Picked["Picked"] = False

    df_Order_Picked["Picked"] = df_Order_Picked["Picked"].map({True: "Picked", False: "No"}).fillna("No")


    # Picked qty per part
    picked_parts = (
        df_Order_Picked.loc[df_Order_Picked["Picked"].eq("Picked")]
        .groupby("Item", as_index=False)["Qty"].sum()
        .rename(columns={"Item": "Part_Number", "Qty": "Picked_Qty"})
    )

    # Inventory merge
    inv_plus = inventory_df.merge(picked_parts, on="Part_Number", how="left")
    for c in ["On Hand", "On Sales Order", "On PO", "Picked_Qty"]:
        if c in inv_plus.columns:
            inv_plus[c] = pd.to_numeric(inv_plus[c], errors="coerce").fillna(0)

    structured_df = df_Order_Picked.merge(
        inv_plus, how="left", left_on="Item", right_on="Part_Number"
    )
    structured_df["Qty"] = pd.to_numeric(structured_df["Qty"], errors="coerce")
    structured_df = structured_df.dropna(subset=["Qty"])

    # Lead Time + assigned totals per Item
    structured_df["Lead Time"] = pd.to_datetime(structured_df["Lead Time"], errors="coerce").dt.floor("D")
    # Convert to datetime first (already in your code)
    structured_df["Lead Time"] = pd.to_datetime(structured_df["Lead Time"], errors="coerce").dt.floor("D")

    # --- Fix dummy dates: move them to 2099 equivalents ---
    mask_july4  = (structured_df["Lead Time"].dt.month.eq(7))  & (structured_df["Lead Time"].dt.day.eq(4))
    mask_dec31  = (structured_df["Lead Time"].dt.month.eq(12)) & (structured_df["Lead Time"].dt.day.eq(31))

    structured_df.loc[mask_july4, "Lead Time"] = pd.Timestamp("2099-07-04")
    structured_df.loc[mask_dec31, "Lead Time"] = pd.Timestamp("2099-12-31")
    assigned_mask = ~(
        (structured_df["Lead Time"].dt.month.eq(7)  & structured_df["Lead Time"].dt.day.eq(4)) |
        (structured_df["Lead Time"].dt.month.eq(12) & structured_df["Lead Time"].dt.day.eq(31))
    )
    assigned_total = structured_df["Qty"].where(assigned_mask, 0).groupby(structured_df["Item"]).transform("sum")
    structured_df["Assigned Q'ty"] = assigned_total
    structured_df["On Hand - WIP"] = structured_df["On Hand"] - structured_df.get("Picked_Qty", 0)


    # Filter pods that have been locked to SO  
    # ['Name'] represents the Customer, ['Source Name'] represents the Vendor
    filtered = df_pod[~df_pod['Name'].isin([
    'Neousys Technology Incorp.',
    'Amazon',
    'Newegg Business, Inc.',
    'Newegg.com',
    'Kontron America, Inc.',
    'Provantage LLC',
    'SMART Modular Technologies, Inc.',
    'Spectrum Sourcing',
    'Arrow Electronics, Inc.',
    'ASI Computer Technologies, Inc.',
    'B&H',
    'PhyTools',
    'Mouser Electronics',
    'Genoedge Corporation DBA SabrePC.COM',
    'CoastIPC, Inc.',
    'Industrial PC, Inc.',

])]
    result = (
        filtered.groupby('Item', as_index=False)['Qty(+)']
        .sum()
    )
    lookup = (
        result[['Item', 'Qty(+)']]
        .drop_duplicates(subset=['Item'])         # ensures uniqueness
        .set_index('Item')['Qty(+)'] # Series: index = part_number
    )
    structured_df['Pre-installed PO'] = structured_df['Item'].map(lookup).fillna(0)
    structured_df["Available + Pre-installed PO"] = structured_df["Available"] + structured_df['Pre-installed PO']

    ## Recommend Restocking QTY
    # Ensure numeric types and fill NaNs
    structured_df['Reorder Pt (Min)'] = pd.to_numeric(structured_df['Reorder Pt (Min)'], errors='coerce').fillna(0)
    structured_df['Available'] = pd.to_numeric(structured_df['Available'], errors='coerce').fillna(0)
    structured_df['On PO'] = pd.to_numeric(structured_df['On PO'], errors='coerce').fillna(0)

    structured_df['Available + On PO'] = structured_df['Available'] + structured_df['On PO']

    # Calculate Restock Qty
    structured_df['Recommended Restock Qty'] = np.ceil(
    np.maximum(0, (4 * structured_df['Sales/Week']) - structured_df['Available'] - structured_df['On PO'])
).astype(int)

    ## Define Component Status
    structured_df["Component_Status"] = np.where((structured_df["Available + Pre-installed PO"] >= 0) & (structured_df["On Hand"] > 0), "Available", "Shortage") #Available or Shortage   
    structured_df["Qty(+)"] = "0"
    structured_df['Pre/Bare'] = "Out"

    structured_df.rename(columns={"SO Entry Date":"Order Date", "Customer": "Name", "Lead Time": "Ship Date", "Customer PO": "P. O. #", "Qty": "Qty(-)", "SO Status": "SO_Status" },inplace=True)
    for col in ["Order Date", "Ship Date"]:
        structured_df[col] = pd.to_datetime(structured_df[col], errors="coerce").dt.date

    return structured_df, final_sales_order


# No Assigned SO Summary
def save_not_assigned_so(
    df: pd.DataFrame,
    output_path: str = "Not_assigned_SO.xlsx",
    highlight_col: str = "Recommended Restock Qty",
    band_by_col: str = "QB Num",
    shortage_col: str = "Component_Status",
    shortage_value: str = "Shortage",
    column_widths: dict | None = None,
) -> dict:
    """
    Save `df` to `output_path`, replacing the first sheet, then apply formatting:
      - Freeze header row
      - Band rows by changes in `band_by_col`
      - Red font for rows where `shortage_col == shortage_value`
      - Highlight `highlight_col` cells > 0 (yellow)
      - Set column widths
      - Rename sheet to today's date (YYYY-MM-DD)
    Returns a summary dict.
    """

    # ---------- defaults ----------
    if column_widths is None:
        column_widths = {
            'Order Date': 15,
            "Item": 30,
            "Name": 25,
            "P. O. #": 15,
            "QB Num": 15,
            "Qty(-)": 10,
            "Available": 15,
            'Available + Pre-installed PO': 25,
            'On Hand - WIP': 20,
            'Reorder Pt (Min)': 15,
            'Recommended Restock Qty': 20,
            'On Sales Order': 15,
        }

    # ---------- ensure workbook exists; if not, create with a temp sheet ----------
    if not os.path.exists(output_path):
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # create a placeholder first sheet (name will be replaced)
            df.to_excel(writer, sheet_name="Sheet1", index=False)

    # ---------- find current first sheet name ----------
    _wb = load_workbook(output_path)
    first_sheet_name = _wb.sheetnames[0]
    _wb.close()

    # ---------- write df to first sheet (replace) ----------
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=first_sheet_name, index=False)

    # ---------- open workbook for styling ----------
    wb = load_workbook(output_path)
    ws = wb.worksheets[0]  # first sheet

    # Freeze first row
    ws.freeze_panes = "A2"

    # Build header map
    col_map: dict[str, int] = {}
    band_col_idx = None
    shortage_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        header = cell.value
        col_map[header] = idx
        if header == band_by_col:
            band_col_idx = idx
        if header == shortage_col:
            shortage_col_idx = idx

    # Fills / fonts / align
    gray_fill    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    yellow_fill  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="00FF0000")  # <-- ARGB, full alpha channel
    center_align = Alignment(horizontal="center", vertical="center")

    # ---------- banding + shortage red ----------
    if band_col_idx is not None and shortage_col_idx is not None:
        current_key = None
        fill_toggle = False

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            band_cell = row[band_col_idx - 1]
            status_cell = row[shortage_col_idx - 1]

            if band_cell.value != current_key:
                current_key = band_cell.value
                fill_toggle = not fill_toggle

            row_fill = gray_fill if fill_toggle else white_fill
            for c in row:
                c.fill = row_fill

            if status_cell.value == shortage_value:
                for c in row:
                    c.font = red_font

    # ---------- highlight target column (cells > 0) ----------
    if highlight_col in col_map:
        h_idx = col_map[highlight_col]
        for (cell,) in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=h_idx, max_col=h_idx):
            try:
                val = float(cell.value)
            except (TypeError, ValueError):
                val = 0.0
            if val > 0:
                cell.fill = yellow_fill  # override banding for this cell

    # ---------- set column widths ----------
    for name, width in column_widths.items():
        if name in col_map:
            letter = ws.cell(row=1, column=col_map[name]).column_letter
            ws.column_dimensions[letter].width = width

    # ---------- center-align a few common numeric columns ----------
    for name in ["Qty", "Available + Pre-installed PO", "Available"]:
        if name in col_map:
            idx = col_map[name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    cell.alignment = center_align

    # ---------- rename sheet to today's date ----------
    today_str = datetime.today().strftime("%Y-%m-%d")
    ws.title = today_str


    # ---------- save ----------
    wb.save(output_path)

    # ---------- summary ----------
    unique_wo = df[band_by_col].nunique() if band_by_col in df.columns else 0
    return {
        "Number of unassigned WOs:": unique_wo,
        "sheet_name": today_str,
    }

# --------------------
# Runner
# --------------------
def main():
    # 1) Extract inputs
    so_raw, inv_raw, ship, df_pod = extract_inputs()

    # 2) External sources
    word_files_df = fetch_word_files_df("http://192.168.60.133:5001/api/word-files")

    # 3) Transform
    so_full = transform_sales_order(so_raw)   # <-- keep the full frame (with Qty, Lead Time, etc.)
    inv = transform_inventory(inv_raw)
    df_pod = transform_pod(df_pod)
    ship = transform_shipping(ship)

    # 4) PDF orders from Supabase -> two columns ["WO","Product Number"]
    pdf_orders_df = fetch_pdf_orders_df_from_supabase(DATABASE_DSN)

    # 5) Build structured_df
    structured, final_sales_order = build_structured_df(so_full, word_files_df, inv, pdf_orders_df, df_pod)

    # 6) Wirte to Not_assigned_SO.xlsx
    ERP_df= structured[['Order Date', "Name", "P. O. #", "QB Num", "Item", 'Qty(-)', 
                              "Available + Pre-installed PO", 'Available', "Assigned Q'ty", 'On Hand - WIP', 'On Hand', 'On Sales Order', 'On PO', 'Reorder Pt (Min)', 'Available + On PO', 'Sales/Week', 'Recommended Restock Qty', 'Ship Date', 'Picked', 'Component_Status']].copy()
    ERP_df["Ship Date"] = pd.to_datetime(ERP_df["Ship Date"], errors="coerce")
    assigned_mask = (
    (ERP_df["Ship Date"].dt.month.eq(7)  & ERP_df["Ship Date"].dt.day.eq(4)) |
    (ERP_df["Ship Date"].dt.month.eq(12) & ERP_df["Ship Date"].dt.day.eq(31))
    )
    Not_assgned_SO = ERP_df[assigned_mask].copy()
    # Summary
    summary = save_not_assigned_so(
    Not_assgned_SO,
    output_path="Not_assigned_SO.xlsx",
    highlight_col="Recommended Restock Qty",
    band_by_col="QB Num",
    shortage_col="Component_Status",
    shortage_value="Shortage",
)
    print(summary)


    # 6) Load to Supabase
    write_inventory_status(inv, table=TBL_INVENTORY, schema=DB_SCHEMA)
    write_sales_order(so_full, table=TBL_SALES_ORDER, schema=DB_SCHEMA)
    write_structured(structured, table=TBL_STRUCTURED, schema=DB_SCHEMA)
    write_pod(df_pod, table=TBL_POD, schema=DB_SCHEMA)
    write_Shipping_Schedule(ship, table=TBL_Shipping, schema=DB_SCHEMA)

    print(f"✅ Loaded:{DB_SCHEMA}.{TBL_SALES_ORDER} rows={len(so_full)}; {DB_SCHEMA}.{TBL_INVENTORY} rows={len(inv)}; {DB_SCHEMA}.{TBL_STRUCTURED} rows={len(structured)}; {DB_SCHEMA}.{TBL_POD} rows={len(df_pod)}; {DB_SCHEMA}.{TBL_Shipping} rows={len(ship)}")


    # 7)Upload final_sales_order to Google Sheets
    if not final_sales_order.empty:
        write_final_sales_order_to_gsheet(
            final_sales_order.assign(**{
                # Optional: make Lead Time a date for the sheet only
                "Lead Time": pd.to_datetime(final_sales_order["Lead Time"], errors="coerce").dt.date
            }),
            spreadsheet_name="PDF_WO",
            worksheet_name="Open Sales Order",
            cred_path=r"C:\Users\Admin\Downloads\pdfwo-466115-734096e1cef8.json",
        )
    else:
        logging.info("No Open Sales Order rows to write to Google Sheets.")


if __name__ == "__main__":
    try:
        logging.info("Running: %s", __file__)
        main()
        logging.info("Done.")
    except Exception as e:
        logging.error("FATAL: %s", e)
        traceback.print_exc()
        sys.exit(1)



# TODO

# Automatically update the Word-files API

# Add validate= to merges and require_cols() checks.

# Replace np.inf with a large int sentinel for stable sorting.

# Add HTTP retries for the Word-files API.

# Gate external side-effects behind flags/env so local dev never blocks.
