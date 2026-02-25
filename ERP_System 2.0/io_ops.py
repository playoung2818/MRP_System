from __future__ import annotations
import os, json, requests, pandas as pd, numpy as np, tempfile, subprocess, uuid
from io import BytesIO
from pathlib import Path
try:
    import gspread
    from gspread_dataframe import set_with_dataframe
    from oauth2client.service_account import ServiceAccountCredentials
except ImportError:
    gspread = None
    set_with_dataframe = None
    ServiceAccountCredentials = None

from core import normalize_wo_number
from db_config import get_engine

from config import SALES_ORDER_FILE, WAREHOUSE_INV_FILE, SHIPPING_SCHEDULE_FILE, POD_FILE

# ---------- DB engine ----------
def engine():
    return get_engine()

def _copy_via_powershell(src: str, dst: str) -> None:
    safe_src = src.replace("'", "''")
    safe_dst = dst.replace("'", "''")
    ps_cmd = f"Copy-Item -LiteralPath '{safe_src}' -Destination '{safe_dst}' -Force"
    res = subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_cmd],
        capture_output=True,
        text=True,
        check=False,
    )
    if res.returncode != 0:
        raise OSError(
            f"PowerShell copy failed for Excel fallback: {src} -> {dst}; "
            f"stderr={res.stderr.strip()}"
        )

def read_excel_safe(path: str | Path, **kwargs) -> pd.DataFrame:
    """
    Read Excel robustly on environments where direct .xlsx open may fail.
    Falls back to a PowerShell copy + in-memory read.
    """
    path_str = str(path)
    try:
        return pd.read_excel(path_str, **kwargs)
    except OSError as e:
        if e.errno != 22:
            raise

    tmp_zip = os.path.join(tempfile.gettempdir(), f"excel_fallback_{uuid.uuid4().hex}.zip")
    try:
        _copy_via_powershell(path_str, tmp_zip)
        with open(tmp_zip, "rb") as f:
            data = f.read()
        return pd.read_excel(BytesIO(data), **kwargs)
    finally:
        try:
            if os.path.exists(tmp_zip):
                os.remove(tmp_zip)
        except Exception:
            pass

# ---------- Extract ----------
def extract_inputs():
    # Force str paths for Windows/OneDrive oddities; explicit engine to avoid parser quirks
    df_sales_order       = pd.read_csv(str(SALES_ORDER_FILE), encoding="ISO-8859-1", engine="python")
    inventory_df         = pd.read_csv(str(WAREHOUSE_INV_FILE))
    df_shipping_schedule = read_excel_safe(SHIPPING_SCHEDULE_FILE)
    df_pod               = pd.read_csv(str(POD_FILE), encoding="ISO-8859-1", engine="python")
    return df_sales_order, inventory_df, df_shipping_schedule, df_pod

def fetch_word_files_df(api_url: str | list[str] | tuple[str, ...]) -> pd.DataFrame:
    urls = [api_url] if isinstance(api_url, str) else list(api_url)
    wf = pd.DataFrame(columns=["file_name", "order_id", "status"])
    for url in urls:
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            data = r.json()
            wf = pd.DataFrame(data.get("word_files", []))
            break
        except Exception:
            continue
    if "order_id" in wf.columns:
        wf = wf.rename(columns={"order_id":"WO_Number"})
    wf["WO_Number"] = wf["WO_Number"].astype(str).apply(normalize_wo_number)
    return wf

def fetch_pdf_orders_df_from_supabase() -> pd.DataFrame:
    """Return columns ['WO','Product Number','Consigned'] built from pdf_file_log.extracted_data JSON."""
    eng = get_engine()
    rows = pd.read_sql('SELECT order_id, extracted_data FROM public.pdf_file_log', eng)

    def rows_from_json(extracted_data, order_id=""):
        if isinstance(extracted_data, str):
            try:
                extracted_data = json.loads(extracted_data)
            except Exception:
                extracted_data = {}
        data = extracted_data or {}
        wo = normalize_wo_number(data.get("wo", order_id))
        consigned = data.get("Consigned")
        if consigned is None:
            consigned = data.get("consigned")
        consigned = bool(consigned) if consigned is not None else False
        items = data.get("items") or []
        if not items:
            return [{"WO": wo, "Product Number": "", "Consigned": consigned}]
        out = []
        for it in items:
            pn = it.get("product_number") or it.get("part_number") or it.get("product") or it.get("part") or ""
            out.append({"WO": wo, "Product Number": pn, "Consigned": consigned})
        return out

    all_rows = []
    for _, r in rows.iterrows():
        all_rows.extend(rows_from_json(r.get("extracted_data"), r.get("order_id")))
    return pd.DataFrame(all_rows, columns=["WO", "Product Number", "Consigned"])

# ---------- Load (DB) ----------
def write_to_db(df: pd.DataFrame, schema: str, table: str):
    if df is None:
        return

    out = df.copy()
    # Normalize nested objects so DB adapters do not choke on dict/list payloads.
    for c in out.columns:
        if out[c].dtype == "object":
            out[c] = out[c].map(
                lambda v: json.dumps(v, ensure_ascii=False)
                if isinstance(v, (dict, list, tuple, set))
                else v
            )

    # Keep SQLAlchemy parameter payload stable.
    out = out.where(pd.notna(out), None)

    eng = engine()
    try:
        out.to_sql(
            table,
            eng,
            schema=schema,
            if_exists="replace",
            index=False,
            method="multi",
            chunksize=2_000,
        )
    except Exception as exc:
        msg = str(exc)
        # Fallback for SQLAlchemy bulk bind failures (e.g. e3q8 / bind parameter errors).
        if ("sqlalche.me/e/20/e3q8" in msg) or ("bind parameter" in msg.lower()):
            out.to_sql(
                table,
                eng,
                schema=schema,
                if_exists="replace",
                index=False,
                method=None,
                chunksize=500,
            )
        else:
            raise

# ---------- Google Sheets ----------
def write_final_sales_order_to_gsheet(df: pd.DataFrame, *,
    spreadsheet_name: str = "PDF_WO",
    worksheet_name: str = "Open Sales Order",
    cred_path: str = r"C:\Users\Admin\OneDrive - neousys-tech\Desktop\Python\pdfwo-466115-734096e1cef8.json",
    consigned_wos: set[str] | None = None,
):
    if gspread is None or set_with_dataframe is None or ServiceAccountCredentials is None:
        raise ImportError(
            "Google Sheets dependencies are missing. Install: "
            "pip install gspread gspread-dataframe oauth2client"
        )
    scope = ["https://spreadsheets.google.com/feeds","https://www.googleapis.com/auth/drive"]
    cred_path_use = cred_path
    temp_cred_path = None
    try:
        temp_cred_path = os.path.join(tempfile.gettempdir(), f"gsheet_cred_{uuid.uuid4().hex}.json")
        _copy_via_powershell(cred_path, temp_cred_path)
        cred_path_use = temp_cred_path
    except Exception:
        cred_path_use = cred_path
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(cred_path_use, scope)
        client = gspread.authorize(creds)
        sh = client.open(spreadsheet_name)
        try:
            ws = sh.worksheet(worksheet_name); ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=worksheet_name, rows=100, cols=26)
        set_with_dataframe(ws, df, include_index=False, include_column_header=True, resize=True)
        try: ws.freeze(rows=1)
        except Exception: pass
        if consigned_wos:
            try:
                from gspread.utils import rowcol_to_a1
                qb_idx = None
                for idx, col in enumerate(df.columns, 1):
                    if str(col).strip() == "QB Num":
                        qb_idx = idx
                        break
                if qb_idx is not None:
                    max_col = len(df.columns)
                    red_fill = {"backgroundColor": {"red": 1.0, "green": 0.8, "blue": 0.8}}
                    for i, qb in enumerate(df["QB Num"].astype(str), start=2):
                        if qb in consigned_wos:
                            start = rowcol_to_a1(i, 1)
                            end = rowcol_to_a1(i, max_col)
                            ws.format(f"{start}:{end}", red_fill)
            except Exception:
                pass
        print(f"Wrote {len(df)} rows to Google Sheet -> {spreadsheet_name} / {worksheet_name}")
    finally:
        if temp_cred_path:
            try:
                if os.path.exists(temp_cred_path):
                    os.remove(temp_cred_path)
            except Exception:
                pass


def merge_open_sales_order_to_allocation_reference_gsheet(
    df: pd.DataFrame,
    *,
    spreadsheet_name: str = "PDF_WO",
    worksheet_name: str = "allocation_reference",
    diff_worksheet_name: str = "allocation_reference_diff",
    cred_path: str = r"C:\Users\Admin\OneDrive - neousys-tech\Desktop\Python\pdfwo-466115-734096e1cef8.json",
) -> dict:
    """
    Merge current Open Sales Order rows into allocation_reference sheet.
    Preserves manual input columns: Pre/Bare, POD.
    Writes a row-level/column-level diff view to diff worksheet.
    """
    if gspread is None or set_with_dataframe is None or ServiceAccountCredentials is None:
        raise ImportError(
            "Google Sheets dependencies are missing. Install: "
            "pip install gspread gspread-dataframe oauth2client"
        )

    ref_cols = ["Customer", "Customer PO", "QB Num", "Item", "Qty", "Lead Time", "Pre/Bare", "POD"]
    sys_cols = ["Customer", "Customer PO", "QB Num", "Item", "Qty", "Lead Time"]
    input_cols = ["Pre/Bare", "POD"]

    def _col(series_df: pd.DataFrame, *candidates: str) -> pd.Series:
        for c in candidates:
            if c in series_df.columns:
                return series_df[c]
        return pd.Series([""] * len(series_df), index=series_df.index, dtype="object")

    def _norm_ref_frame(src: pd.DataFrame) -> pd.DataFrame:
        out = pd.DataFrame(index=src.index)
        out["Customer"] = _col(src, "Customer", "Name").astype(str).str.strip()
        out["Customer PO"] = _col(src, "Customer PO", "P. O. #").astype(str).str.strip()
        out["QB Num"] = _col(src, "QB Num").astype(str).str.strip()
        out["Item"] = _col(src, "Item").astype(str).str.strip()
        out["Qty"] = pd.to_numeric(_col(src, "Qty", "Qty(-)"), errors="coerce").fillna(0.0)
        lead = pd.to_datetime(_col(src, "Lead Time", "Ship Date"), errors="coerce")
        out["Lead Time"] = lead.dt.strftime("%Y-%m-%d").fillna("")
        out["Pre/Bare"] = _col(src, "Pre/Bare").astype(str).replace("nan", "").str.strip()
        out["POD"] = _col(src, "POD").astype(str).replace("nan", "").str.strip()
        return out[ref_cols].copy()

    def _add_row_key(frame: pd.DataFrame) -> pd.DataFrame:
        x = frame.copy()
        x["__ord"] = (
            x["Customer PO"].astype(str) + "|" + x["Lead Time"].astype(str) + "|" + x["Qty"].astype(str) + "|" + x["Customer"].astype(str)
        )
        x = x.sort_values(["QB Num", "Item", "__ord"]).reset_index(drop=True)
        x["__occ"] = x.groupby(["QB Num", "Item"]).cumcount()
        x["__row_key"] = x["QB Num"].astype(str) + "||" + x["Item"].astype(str) + "||" + x["__occ"].astype(str)
        x.drop(columns=["__ord"], inplace=True)
        return x

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    cred_path_use = cred_path
    temp_cred_path = None
    try:
        temp_cred_path = os.path.join(tempfile.gettempdir(), f"gsheet_cred_{uuid.uuid4().hex}.json")
        _copy_via_powershell(cred_path, temp_cred_path)
        cred_path_use = temp_cred_path
    except Exception:
        cred_path_use = cred_path

    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(cred_path_use, scope)
        client = gspread.authorize(creds)
        sh = client.open(spreadsheet_name)

        # Existing reference sheet (if any)
        try:
            ws_ref = sh.worksheet(worksheet_name)
            old_records = ws_ref.get_all_records()
            old_df = pd.DataFrame(old_records) if old_records else pd.DataFrame(columns=ref_cols)
        except gspread.exceptions.WorksheetNotFound:
            ws_ref = sh.add_worksheet(title=worksheet_name, rows=500, cols=20)
            old_df = pd.DataFrame(columns=ref_cols)

        # Normalize both sides
        new_df = _norm_ref_frame(df.copy())
        old_df = _norm_ref_frame(old_df.copy()) if not old_df.empty else pd.DataFrame(columns=ref_cols)

        old_k = _add_row_key(old_df)
        new_k = _add_row_key(new_df)

        # Preserve user input columns from existing sheet by row_key
        keep = old_k[["__row_key"] + input_cols].copy() if not old_k.empty else pd.DataFrame(columns=["__row_key"] + input_cols)
        merged = new_k.merge(keep, on="__row_key", how="left", suffixes=("", "_old"))
        for c in input_cols:
            old_c = f"{c}_old"
            if old_c in merged.columns:
                merged[c] = merged[old_c].where(merged[old_c].astype(str).str.strip().ne(""), merged[c])
                merged.drop(columns=[old_c], inplace=True)
            merged[c] = merged[c].fillna("").astype(str).str.strip()

        out_df = merged[ref_cols].copy()

        # Build diff view
        now_utc = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        diff_rows: list[dict] = []
        old_map = old_k.set_index("__row_key") if not old_k.empty else pd.DataFrame(columns=old_k.columns).set_index(pd.Index([], name="__row_key"))
        new_map = merged.set_index("__row_key") if not merged.empty else pd.DataFrame(columns=merged.columns).set_index(pd.Index([], name="__row_key"))

        old_keys = set(old_map.index.tolist())
        new_keys = set(new_map.index.tolist())

        for k in sorted(new_keys - old_keys):
            r = new_map.loc[k]
            diff_rows.append(
                {
                    "generated_at_utc": now_utc,
                    "change_type": "ADDED",
                    "row_key": k,
                    "QB Num": r.get("QB Num", ""),
                    "Item": r.get("Item", ""),
                    "column": "",
                    "old_value": "",
                    "new_value": "new row",
                }
            )
        for k in sorted(old_keys - new_keys):
            r = old_map.loc[k]
            diff_rows.append(
                {
                    "generated_at_utc": now_utc,
                    "change_type": "REMOVED",
                    "row_key": k,
                    "QB Num": r.get("QB Num", ""),
                    "Item": r.get("Item", ""),
                    "column": "",
                    "old_value": "old row",
                    "new_value": "",
                }
            )
        for k in sorted(new_keys & old_keys):
            r_old = old_map.loc[k]
            r_new = new_map.loc[k]
            for c in ref_cols:
                ov = "" if pd.isna(r_old.get(c, "")) else str(r_old.get(c, ""))
                nv = "" if pd.isna(r_new.get(c, "")) else str(r_new.get(c, ""))
                if ov != nv:
                    diff_rows.append(
                        {
                            "generated_at_utc": now_utc,
                            "change_type": "CHANGED",
                            "row_key": k,
                            "QB Num": r_new.get("QB Num", ""),
                            "Item": r_new.get("Item", ""),
                            "column": c,
                            "old_value": ov,
                            "new_value": nv,
                        }
                    )

        diff_df = pd.DataFrame(
            diff_rows,
            columns=["generated_at_utc", "change_type", "row_key", "QB Num", "Item", "column", "old_value", "new_value"],
        )
        if diff_df.empty:
            diff_df = pd.DataFrame(
                [
                    {
                        "generated_at_utc": now_utc,
                        "change_type": "NO_CHANGE",
                        "row_key": "",
                        "QB Num": "",
                        "Item": "",
                        "column": "",
                        "old_value": "",
                        "new_value": "",
                    }
                ]
            )

        ws_ref.clear()
        set_with_dataframe(ws_ref, out_df, include_index=False, include_column_header=True, resize=True)
        try:
            ws_ref.freeze(rows=1)
        except Exception:
            pass

        try:
            ws_diff = sh.worksheet(diff_worksheet_name)
            ws_diff.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws_diff = sh.add_worksheet(title=diff_worksheet_name, rows=500, cols=20)
        set_with_dataframe(ws_diff, diff_df, include_index=False, include_column_header=True, resize=True)
        try:
            ws_diff.freeze(rows=1)
        except Exception:
            pass

        summary = {
            "rows_written": int(len(out_df)),
            "added": int((diff_df["change_type"] == "ADDED").sum()),
            "removed": int((diff_df["change_type"] == "REMOVED").sum()),
            "changed_cells": int((diff_df["change_type"] == "CHANGED").sum()),
        }
        print(
            "Merged Open Sales Order to Google Sheet -> "
            f"{spreadsheet_name} / {worksheet_name}; diff -> {diff_worksheet_name}; summary={summary}"
        )
        return summary
    finally:
        if temp_cred_path:
            try:
                if os.path.exists(temp_cred_path):
                    os.remove(temp_cred_path)
            except Exception:
                pass

# ---------- Excel styling helper functions ----------
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def save_not_assigned_so(
    df: pd.DataFrame,
    output_path: str = "Not_assigned_SO.xlsx",
    highlight_cols: str | list[str] | None = None,
    band_by_col: str = "QB Num",
    shortage_col: str = "Component_Status",
    shortage_value: str = "Shortage",
    column_widths: dict | None = None,
    pod_watchlist_df: pd.DataFrame | None = None,
    pod_watchlist_sheet: str = "POD-Wachlist",
) -> dict:
    """
    Save `df` to `output_path`, replacing the first sheet, then apply formatting:
      - Freeze header row
      - Band rows by changes in `band_by_col`
      - Red font for rows where `shortage_col == shortage_value`
      - Highlight `highlight_col` cells > 0 (yellow)
      - Set column widths
      - Rename sheet to today's date (YYYY-MM-DD)
    Returns a summary dict.
    """

    # ---------- defaults ----------
    if column_widths is None:
        column_widths = {
            'Order Date': 15,
            "Item": 30,
            "Name": 25,
            "P. O. #": 5,
            "QB Num": 15,
            "Qty(-)": 10,
            "Available": 15,
            'Available + Pre-installed PO': 25,
            'On Hand - WIP': 17,
            'Recommended Restock Qty': 25,
            'On Sales Order': 15,
            'Available + On PO': 20,
            "Assigned Q'ty": 15,
            "Sales/Week": 15, 
        }

    output_path = str(output_path)
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    use_xlsx_fallback = output_path.lower().endswith(".xlsx")

    if use_xlsx_fallback:
        wb = None
        if os.path.exists(output_path):
            temp_in = os.path.join(tempfile.gettempdir(), f"not_assigned_in_{uuid.uuid4().hex}.zip")
            try:
                _copy_via_powershell(output_path, temp_in)
                with open(temp_in, "rb") as f:
                    wb = load_workbook(BytesIO(f.read()))
            except Exception:
                wb = None
            finally:
                try:
                    if os.path.exists(temp_in):
                        os.remove(temp_in)
                except Exception:
                    pass
        if wb is None:
            wb = Workbook()
        first_sheet_name = wb.sheetnames[0] if wb.sheetnames else "Sheet1"
        if first_sheet_name in wb.sheetnames:
            del wb[first_sheet_name]
        ws = wb.create_sheet(title=first_sheet_name, index=0)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        if pod_watchlist_df is not None:
            if pod_watchlist_sheet in wb.sheetnames:
                del wb[pod_watchlist_sheet]
            ws_pod = wb.create_sheet(title=pod_watchlist_sheet)
            for row in dataframe_to_rows(pod_watchlist_df, index=False, header=True):
                ws_pod.append(row)
    else:
        # ---------- ensure workbook exists; if not, create with a temp sheet ----------
        if not os.path.exists(output_path):
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # create a placeholder first sheet (name will be replaced)
                df.to_excel(writer, sheet_name="Sheet1", index=False)

        # ---------- find current first sheet name ----------
        _wb = load_workbook(output_path)
        first_sheet_name = _wb.sheetnames[0]
        _wb.close()

        # ---------- write df to first sheet (replace) ----------
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=first_sheet_name, index=False)
            if pod_watchlist_df is not None:
                pod_watchlist_df.to_excel(writer, sheet_name=pod_watchlist_sheet, index=False)

        # ---------- open workbook for styling ----------
        wb = load_workbook(output_path)
        ws = wb.worksheets[0]  # first sheet

    # Freeze first row
    ws.freeze_panes = "A2"

    # Build header map
    col_map: dict[str, int] = {}
    band_col_idx = None
    shortage_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        header = cell.value
        col_map[header] = idx
        if header == band_by_col:
            band_col_idx = idx
        if header == shortage_col:
            shortage_col_idx = idx

    # Fills / fonts / align
    gray_fill    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    yellow_fill  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="00FF0000")  # <-- ARGB, full alpha channel
    center_align = Alignment(horizontal="center", vertical="center")

    # ---------- banding + shortage red ----------
    if band_col_idx is not None and shortage_col_idx is not None:
        current_key = None
        fill_toggle = False

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            band_cell = row[band_col_idx - 1]
            status_cell = row[shortage_col_idx - 1]

            if band_cell.value != current_key:
                current_key = band_cell.value
                fill_toggle = not fill_toggle

            row_fill = gray_fill if fill_toggle else white_fill
            for c in row:
                c.fill = row_fill

            if status_cell.value == shortage_value:
                for c in row:
                    c.font = red_font

    # ---------- warn rows where Sales/Week > Available + On PO ----------
    sales_idx = col_map.get("Sales/Week")
    avail_on_po_idx = col_map.get("Available + On PO")
    if sales_idx and avail_on_po_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            try:
                sales_val = float(row[sales_idx - 1].value)
            except (TypeError, ValueError):
                sales_val = 0.0
            try:
                avail_val = float(row[avail_on_po_idx - 1].value)
            except (TypeError, ValueError):
                avail_val = 0.0
            if sales_val > avail_val:
                target_cell = row[avail_on_po_idx - 1]
                target_cell.fill = yellow_fill

    # ---------- highlight target column (cells > 0) ----------
    if highlight_cols is None:
        highlight_cols = []
    if isinstance(highlight_cols, str):
        highlight_cols = [highlight_cols]
    for highlight_col in highlight_cols:
        if highlight_col in col_map:
            h_idx = col_map[highlight_col]
            for (cell,) in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=h_idx, max_col=h_idx):
                try:
                    val = float(cell.value)
                except (TypeError, ValueError):
                    val = 0.0
                if val > 0:
                    cell.fill = yellow_fill  # override banding for this cell

    # ---------- set column widths ----------
    for name, width in column_widths.items():
        if name in col_map:
            letter = ws.cell(row=1, column=col_map[name]).column_letter
            ws.column_dimensions[letter].width = width

    # ---------- center-align a few common numeric columns ----------
    for name in ["Qty(-)", "Available + Pre-installed PO", "Available", "Available + On PO", "Sales/Week", "Recommended Restock Qty"]:
        if name in col_map:
            idx = col_map[name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    cell.alignment = center_align

    # ---------- rename sheet to today's date ----------
    today_str = datetime.today().strftime("%Y-%m-%d")
    ws.title = today_str

    # ---------- freeze header on POD watchlist sheet ----------
    if pod_watchlist_df is not None and pod_watchlist_sheet in wb.sheetnames:
        ws_pod = wb[pod_watchlist_sheet]
        ws_pod.freeze_panes = "A2"


    # ---------- save ----------
    if use_xlsx_fallback:
        out_bytes = BytesIO()
        wb.save(out_bytes)
        temp_out = os.path.join(tempfile.gettempdir(), f"not_assigned_out_{uuid.uuid4().hex}.zip")
        with open(temp_out, "wb") as f:
            f.write(out_bytes.getvalue())
        try:
            _copy_via_powershell(temp_out, output_path)
        finally:
            try:
                if os.path.exists(temp_out):
                    os.remove(temp_out)
            except Exception:
                pass
    else:
        wb.save(output_path)

    # ---------- summary ----------
    unique_wo = df[band_by_col].nunique() if band_by_col in df.columns else 0
    return {
        "Number of unassigned WOs:": unique_wo,
        "sheet_name": today_str,
    }
